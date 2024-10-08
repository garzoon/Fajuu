from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file
import json

from ...controller import *
from ...models import Cliente, Factura
import openpyxl
import io

# Decoradores
from ...utils.user_decorators import role_requiered

factura_cliente_scope = Blueprint("factura_cliente_scope", __name__)
PATH_URL_FACT_CLIENTE = "factura/factura_cliente"



@factura_cliente_scope.route('/', methods = ['POST', 'GET'])
@role_requiered([1, 2])
def factura():

    query = """SELECT * FROM facturas WHERE 1=1"""
    parameters = []

    if request.method == 'POST':
        factura_id = request.form.get('factura_id')
        cliente_id = request.form.get('cliente_id')
        factura_fecha = request.form.get('factura_fecha')

        if factura_id:
            query += " AND fact_id LIKE %s"
            parameters.append(f"%{factura_id}%")

        if cliente_id:
            query += " AND clien_copiaid LIKE %s"
            parameters.append(f"%{cliente_id}%")

        if factura_fecha:
            query += " AND fact_fecha_emision LIKE %s"
            parameters.append(f"%{factura_fecha}%")

    factura_list_result = fetch_all(query, parameters)

    list_facturas = []
    
    for factura in factura_list_result:
        factura = Factura(*factura)
        cliente = cliente_select(factura.clien_copiaid)
        item_factura = (
            factura.fact_id, 
            factura.clien_copiaid,
            cliente.clien_nombre,
            factura.fact_fecha_emision
        )
        list_facturas.append(item_factura)

    return render_template(f'{PATH_URL_FACT_CLIENTE}/factura_cliente.html', list_facturas = list_facturas)

@factura_cliente_scope.route('/factura_cliente_delete/<int:id>', methods = ['GET', 'POST'])
@role_requiered([1])
def delete_factura_cliente(id):
    try:
        factura = factura_select(id)
        factura_delete(factura)
        flash(f'Factura de cliente {factura.fact_id} - {factura.clien_copiaid} fue eliminada', "success")
        return redirect(url_for('factura_cliente_scope.factura'))
    
    # Manejo de errores
    except mysql.connector.IntegrityError as ex:
        flash("No se puede eliminar la factura porque está en uso", "warning")
    except mysql.connector.Error as ex:
        flash("Error al intentar eliminar la factura", "warning")
    except Exception as ex:
        flash(f"Error inesperado: {ex}", "warning")
    return redirect(url_for('operador_scope.operador'))

@factura_cliente_scope.route('/factura_cliente_details/<int:id>', methods = ['GET'])
@role_requiered([1, 2])
def details_factura_cliente(id):
    factura = factura_select(id)
    dic_productos = json.loads(factura.fact_detalle_productos)
    cliente = cliente_select(factura.clien_copiaid)
    if factura and cliente:
        return jsonify({
            'id' : factura.fact_id,
            'cliente_id' : factura.clien_copiaid,
            'nombre' : cliente.clien_nombre,
            'fecha' : factura.fact_fecha_emision,
            'dic_productos' : dic_productos,
            'total_precio' : factura.fact_valor_total
        })
    else:
        return jsonify({'Error' : "factura no encontrada"}), 404
    
    
@factura_cliente_scope.route('/factura_cliente_save/<int:id>', methods = ['GET', 'POST'])
@role_requiered([1, 2])
def save_factura_cliente(id):
    factura = factura_select(id)
    cliente = cliente_select(factura.clien_copiaid)
    productos = json.loads(factura.fact_detalle_productos)
    
    list_products = []
    
    for key, producto in productos.items():
        info_producto = producto_select(key)
        info_categoria = get_producto_categoria(info_producto.cate_copiaid)
        list_products.append((
            info_producto.prod_id, 
            info_producto.prod_descripcion, 
            info_categoria.cate_descripcion, 
            producto[1],
            producto[2],
            producto[3]
            ))
    
    libro = openpyxl.Workbook()
    hoja = libro.active
    
    hoja.cell(row=1, column=1, value="Id de Factura")
    hoja.cell(row=1, column=2, value=factura.fact_id)
    
    hoja.cell(row=2, column=1, value="Identificacion de cliente")
    hoja.cell(row=2, column=2, value=cliente.clien_documento)
    
    hoja.cell(row=3, column=1, value="Cliente")
    hoja.cell(row=3, column=2, value=cliente.clien_nombre)
    
    hoja.cell(row=4, column=1, value="Direccion")
    hoja.cell(row=4, column=2, value=f"{cliente.clien_ciudad} - {cliente.clien_direccion}")
    
    hoja.cell(row=5, column=1, value="Correo Electronico")
    hoja.cell(row=5, column=2, value=cliente.clien_email)
    
    hoja.cell(row=6, column=1, value="Numero Telefonico")
    hoja.cell(row=6, column=2, value=cliente.clien_telefono)
    
    hoja.append(("Id producto", "Descripcion", "Categoria", "Cantidad", "Unidad de medida", "Valor"))
    
    for producto in list_products:
        hoja.append(producto)
        
    hoja.append(("Total de Factura", factura.fact_valor_total))
        
    output = io.BytesIO()
    libro.save(output)
    output.seek(0)  # Posicionar el puntero al inicio del archivo

    # Enviar el archivo al cliente
    return send_file(output, as_attachment=True, download_name=f"Factura_{factura.fact_id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    
    